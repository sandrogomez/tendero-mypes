"""
generate_template.py
====================
Generates the Excel accounting template for Chilean small businesses.
Run this script to (re)create templates/contabilidad_pyme.xlsx.

Usage:
    python3 src/generate_template.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent.parent / "templates" / "contabilidad_pyme.xlsx"

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # header background
LIGHT_BLUE  = "BDD7EE"   # sub-header background
LIGHT_GREY  = "F2F2F2"   # alternating row background
WHITE       = "FFFFFF"
GREEN       = "E2EFDA"   # income rows
RED         = "FCE4D6"   # expense rows
ORANGE      = "FCE5CD"   # totals

# ── Reusable style helpers ───────────────────────────────────────────────────

def header_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def normal_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_header_row(ws, row, values, bg=DARK_BLUE):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = header_font()
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = thin_border()


def apply_data_row(ws, row, values, bg=WHITE, bold=False):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = normal_font(bold=bold)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()
    return ws


def set_col_widths(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Sheet 1 – Portada ───────────────────────────────────────────────────────

def build_portada(wb):
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 40
    ws.column_dimensions["A"].width = 60

    ws.merge_cells("A1:A1")
    title = ws["A1"]
    title.value = "PLANTILLA DE CONTABILIDAD – PYME CHILE"
    title.font = Font(bold=True, size=18, color=DARK_BLUE, name="Calibri")
    title.alignment = center()

    info = [
        ("Empresa:",         ""),
        ("RUT empresa:",     ""),
        ("Período:",         "Enero – Diciembre 20XX"),
        ("Responsable:",     ""),
        ("Última actualización:", ""),
    ]
    ws.append([])   # spacer
    for label, value in info:
        ws.append([f"{label} {value}"])

    instructions = [
        "",
        "INSTRUCCIONES DE USO",
        "1. Complete los datos de la empresa en esta hoja.",
        "2. Registre cada transacción en la hoja «Libro Diario».",
        "3. Consulte el resumen mensual en «Ingresos y Egresos».",
        "4. Revise el estado financiero en «Estado de Resultados» y «Balance».",
        "5. La hoja «Plan de Cuentas» contiene el catálogo de cuentas sugerido.",
    ]
    for line in instructions:
        ws.append([line])


# ── Sheet 2 – Plan de Cuentas ───────────────────────────────────────────────

CUENTAS = [
    # (código, nombre, tipo)
    ("1", "ACTIVOS", "Grupo"),
    ("1.1", "Activo Circulante", "Grupo"),
    ("1.1.1", "Caja", "Detalle"),
    ("1.1.2", "Banco", "Detalle"),
    ("1.1.3", "Cuentas por Cobrar", "Detalle"),
    ("1.1.4", "Inventario / Mercaderías", "Detalle"),
    ("1.1.5", "IVA Crédito Fiscal", "Detalle"),
    ("1.2", "Activo Fijo", "Grupo"),
    ("1.2.1", "Muebles y Equipos", "Detalle"),
    ("1.2.2", "Depreciación Acumulada", "Detalle"),
    ("2", "PASIVOS", "Grupo"),
    ("2.1", "Pasivo Circulante", "Grupo"),
    ("2.1.1", "Cuentas por Pagar", "Detalle"),
    ("2.1.2", "IVA Débito Fiscal", "Detalle"),
    ("2.1.3", "Retenciones por Pagar", "Detalle"),
    ("2.2", "Pasivo Largo Plazo", "Grupo"),
    ("2.2.1", "Préstamos Bancarios", "Detalle"),
    ("3", "PATRIMONIO", "Grupo"),
    ("3.1", "Capital", "Detalle"),
    ("3.2", "Resultados del Ejercicio", "Detalle"),
    ("4", "INGRESOS", "Grupo"),
    ("4.1", "Ventas de Productos", "Detalle"),
    ("4.2", "Ventas de Servicios", "Detalle"),
    ("4.3", "Otros Ingresos", "Detalle"),
    ("5", "COSTOS Y GASTOS", "Grupo"),
    ("5.1", "Costo de Ventas", "Detalle"),
    ("5.2", "Remuneraciones", "Detalle"),
    ("5.3", "Arriendos", "Detalle"),
    ("5.4", "Servicios Básicos", "Detalle"),
    ("5.5", "Publicidad y Marketing", "Detalle"),
    ("5.6", "Gastos Financieros", "Detalle"),
    ("5.7", "Depreciación", "Detalle"),
    ("5.8", "Otros Gastos", "Detalle"),
]


def build_plan_cuentas(wb):
    ws = wb.create_sheet("Plan de Cuentas")
    ws.sheet_view.showGridLines = False

    apply_header_row(ws, 1, ["Código", "Nombre de Cuenta", "Tipo"])
    set_col_widths(ws, [12, 40, 14])

    for i, (code, name, tipo) in enumerate(CUENTAS, start=2):
        bg = LIGHT_BLUE if tipo == "Grupo" else WHITE
        bold = tipo == "Grupo"
        apply_data_row(ws, i, [code, name, tipo], bg=bg, bold=bold)


# ── Sheet 3 – Libro Diario ───────────────────────────────────────────────────

def build_libro_diario(wb):
    ws = wb.create_sheet("Libro Diario")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Fecha", "N° Comprobante", "Descripción / Glosa",
               "Código Cuenta", "Nombre Cuenta", "Debe (CLP)", "Haber (CLP)"]
    apply_header_row(ws, 1, headers)
    set_col_widths(ws, [14, 16, 40, 16, 30, 16, 16])

    # Example rows
    examples = [
        ["2024-01-02", "001", "Aporte de capital inicial", "1.1.2", "Banco", 5000000, ""],
        ["2024-01-02", "001", "Aporte de capital inicial", "3.1",   "Capital", "", 5000000],
        ["2024-01-05", "002", "Compra de mercadería", "1.1.4", "Inventario / Mercaderías", 500000, ""],
        ["2024-01-05", "002", "IVA Crédito Fiscal",   "1.1.5", "IVA Crédito Fiscal",        95000, ""],
        ["2024-01-05", "002", "Pago a proveedor",     "1.1.2", "Banco", "", 595000],
    ]
    for row_idx, row_data in enumerate(examples, start=2):
        bg = LIGHT_GREY if row_idx % 2 == 0 else WHITE
        apply_data_row(ws, row_idx, row_data, bg=bg)
        for col in (6, 7):   # format as CLP currency
            cell = ws.cell(row=row_idx, column=col)
            cell.number_format = '#,##0'

    # Totals row template (after example data)
    totals_row = len(examples) + 2
    ws.cell(row=totals_row, column=5, value="TOTALES").font = normal_font(bold=True)
    ws.cell(row=totals_row, column=5).fill = fill(ORANGE)
    for col in (6, 7):
        first = 2
        last  = totals_row - 1
        col_letter = get_column_letter(col)
        cell = ws.cell(row=totals_row, column=col,
                       value=f"=SUM({col_letter}{first}:{col_letter}{last})")
        cell.font = normal_font(bold=True)
        cell.fill = fill(ORANGE)
        cell.number_format = '#,##0'
        cell.border = thin_border()


# ── Sheet 4 – Ingresos y Egresos ────────────────────────────────────────────

MONTHS = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
          "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

INCOME_CATEGORIES = [
    "Ventas de Productos",
    "Ventas de Servicios",
    "Otros Ingresos",
]

EXPENSE_CATEGORIES = [
    "Costo de Ventas",
    "Remuneraciones",
    "Arriendos",
    "Servicios Básicos",
    "Publicidad y Marketing",
    "Gastos Financieros",
    "Depreciación",
    "Otros Gastos",
]


def build_ingresos_egresos(wb):
    ws = wb.create_sheet("Ingresos y Egresos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"

    # Header row: Category | Jan | Feb | … | Dec | Total
    headers = ["Categoría"] + MONTHS + ["TOTAL"]
    apply_header_row(ws, 1, headers)
    set_col_widths(ws, [28] + [12] * 12 + [14])

    current_row = 2

    def section_header(label, bg=DARK_BLUE):
        nonlocal current_row
        ws.merge_cells(f"A{current_row}:N{current_row}")
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = header_font(size=10)
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = thin_border()
        current_row += 1

    def data_section(categories, row_bg):
        nonlocal current_row
        start = current_row
        for cat in categories:
            ws.cell(row=current_row, column=1, value=cat).font = normal_font()
            ws.cell(row=current_row, column=1).border = thin_border()
            ws.cell(row=current_row, column=1).fill = fill(WHITE)
            for col in range(2, 15):
                cell = ws.cell(row=current_row, column=col, value=0)
                cell.number_format = '#,##0'
                cell.fill = fill(row_bg)
                cell.border = thin_border()
            # Total column (=SUM of Jan..Dec)
            total_cell = ws.cell(
                row=current_row, column=14,
                value=f"=SUM(B{current_row}:M{current_row})"
            )
            total_cell.number_format = '#,##0'
            total_cell.fill = fill(ORANGE)
            total_cell.border = thin_border()
            total_cell.font = normal_font(bold=True)
            current_row += 1
        return start, current_row - 1

    def totals_row(label, start, end, bg):
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = normal_font(bold=True)
        cell.fill = fill(bg)
        cell.border = thin_border()
        for col in range(2, 15):
            col_letter = get_column_letter(col)
            c = ws.cell(
                row=current_row, column=col,
                value=f"=SUM({col_letter}{start}:{col_letter}{end})"
            )
            c.number_format = '#,##0'
            c.font = normal_font(bold=True)
            c.fill = fill(bg)
            c.border = thin_border()
        current_row += 1
        return current_row - 1

    # ── INCOME section ──
    section_header("INGRESOS", bg=DARK_BLUE)
    inc_start, inc_end = data_section(INCOME_CATEGORIES, GREEN)
    inc_total_row = totals_row("TOTAL INGRESOS", inc_start, inc_end, LIGHT_BLUE)

    ws.append([])  # spacer
    current_row += 1

    # ── EXPENSE section ──
    section_header("EGRESOS", bg="A50000")
    exp_start, exp_end = data_section(EXPENSE_CATEGORIES, RED)
    exp_total_row = totals_row("TOTAL EGRESOS", exp_start, exp_end, "FCE4D6")

    ws.append([])
    current_row += 1

    # ── NET RESULT ──
    section_header("RESULTADO", bg="375623")
    result_row = current_row
    cell = ws.cell(row=result_row, column=1, value="Resultado del Período")
    cell.font = normal_font(bold=True)
    cell.fill = fill(GREEN)
    cell.border = thin_border()
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        c = ws.cell(
            row=result_row, column=col,
            value=(
                f"={col_letter}{inc_total_row}-{col_letter}{exp_total_row}"
            )
        )
        c.number_format = '#,##0'
        c.font = normal_font(bold=True)
        c.fill = fill(GREEN)
        c.border = thin_border()


# ── Sheet 5 – Estado de Resultados ──────────────────────────────────────────

def build_estado_resultados(wb):
    ws = wb.create_sheet("Estado de Resultados")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [40, 20, 20])

    def title_cell(row, text, bg=DARK_BLUE):
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = header_font()
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

    def row_item(row, label, formula_or_value, bold=False, bg=WHITE):
        ws.cell(row=row, column=1, value=label).font = normal_font(bold=bold)
        ws.cell(row=row, column=1).fill = fill(bg)
        ws.cell(row=row, column=1).border = thin_border()
        cell = ws.cell(row=row, column=2, value=formula_or_value)
        cell.number_format = '#,##0'
        cell.font = normal_font(bold=bold)
        cell.fill = fill(bg)
        cell.border = thin_border()
        note = ws.cell(row=row, column=3, value="")
        note.fill = fill(bg)
        note.border = thin_border()

    title_cell(1, "ESTADO DE RESULTADOS")
    ws.cell(row=2, column=1, value="Empresa:").font = normal_font()
    ws.cell(row=2, column=2, value="").font = normal_font()
    ws.cell(row=3, column=1, value="Período:").font = normal_font()
    ws.cell(row=3, column=2, value="Enero – Diciembre 20XX").font = normal_font()

    title_cell(5, "INGRESOS", bg=DARK_BLUE)
    row_item(6,  "Ventas de Productos",     0)
    row_item(7,  "Ventas de Servicios",     0)
    row_item(8,  "Otros Ingresos",          0)
    row_item(9,  "TOTAL INGRESOS",          "=SUM(B6:B8)", bold=True, bg=LIGHT_BLUE)

    title_cell(11, "COSTOS Y GASTOS", bg="A50000")
    row_item(12, "Costo de Ventas",         0, bg=RED)
    row_item(13, "Remuneraciones",          0, bg=RED)
    row_item(14, "Arriendos",               0, bg=RED)
    row_item(15, "Servicios Básicos",       0, bg=RED)
    row_item(16, "Publicidad y Marketing",  0, bg=RED)
    row_item(17, "Gastos Financieros",      0, bg=RED)
    row_item(18, "Depreciación",            0, bg=RED)
    row_item(19, "Otros Gastos",            0, bg=RED)
    row_item(20, "TOTAL COSTOS Y GASTOS",   "=SUM(B12:B19)", bold=True, bg="FCE4D6")

    title_cell(22, "RESULTADO DEL EJERCICIO", bg="375623")
    row_item(23, "Resultado antes de impuestos", "=B9-B20", bold=True, bg=GREEN)
    row_item(24, "Impuesto a la Renta (PPM/AT)", 0, bg=LIGHT_GREY)
    row_item(25, "RESULTADO NETO",               "=B23-B24", bold=True, bg=GREEN)


# ── Sheet 6 – Balance General ────────────────────────────────────────────────

def build_balance(wb):
    ws = wb.create_sheet("Balance General")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [40, 20, 20])

    def title_cell(row, text, bg=DARK_BLUE):
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = header_font()
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

    def row_item(row, label, value=0, bold=False, bg=WHITE, indent=0):
        label_text = ("    " * indent) + label
        ws.cell(row=row, column=1, value=label_text).font = normal_font(bold=bold)
        ws.cell(row=row, column=1).fill = fill(bg)
        ws.cell(row=row, column=1).border = thin_border()
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = '#,##0'
        cell.font = normal_font(bold=bold)
        cell.fill = fill(bg)
        cell.border = thin_border()
        note = ws.cell(row=row, column=3)
        note.fill = fill(bg)
        note.border = thin_border()

    title_cell(1, "BALANCE GENERAL")
    ws.cell(row=2, column=1, value="Empresa:").font = normal_font()
    ws.cell(row=3, column=1, value="Al:").font = normal_font()

    # ACTIVOS
    title_cell(5, "ACTIVOS")
    row_item(6,  "Activo Circulante",          bold=True, bg=LIGHT_BLUE)
    row_item(7,  "Caja",                        0, indent=1)
    row_item(8,  "Banco",                       0, indent=1)
    row_item(9,  "Cuentas por Cobrar",          0, indent=1)
    row_item(10, "Inventario / Mercaderías",    0, indent=1)
    row_item(11, "IVA Crédito Fiscal",          0, indent=1)
    row_item(12, "Total Activo Circulante",     "=SUM(B7:B11)", bold=True, bg=LIGHT_GREY)
    row_item(13, "Activo Fijo",                 bold=True, bg=LIGHT_BLUE)
    row_item(14, "Muebles y Equipos",           0, indent=1)
    row_item(15, "Depreciación Acumulada",      0, indent=1)
    row_item(16, "Total Activo Fijo",           "=SUM(B14:B15)", bold=True, bg=LIGHT_GREY)
    row_item(17, "TOTAL ACTIVOS",               "=B12+B16", bold=True, bg=LIGHT_BLUE)

    # PASIVOS
    title_cell(19, "PASIVOS Y PATRIMONIO")
    row_item(20, "Pasivo Circulante",           bold=True, bg=RED)
    row_item(21, "Cuentas por Pagar",           0, indent=1)
    row_item(22, "IVA Débito Fiscal",           0, indent=1)
    row_item(23, "Retenciones por Pagar",       0, indent=1)
    row_item(24, "Total Pasivo Circulante",     "=SUM(B21:B23)", bold=True, bg=LIGHT_GREY)
    row_item(25, "Pasivo Largo Plazo",          bold=True, bg=RED)
    row_item(26, "Préstamos Bancarios",         0, indent=1)
    row_item(27, "Total Pasivo Largo Plazo",    "=B26", bold=True, bg=LIGHT_GREY)
    row_item(28, "TOTAL PASIVOS",               "=B24+B27", bold=True, bg=RED)

    row_item(30, "Patrimonio",                  bold=True, bg=GREEN)
    row_item(31, "Capital",                     0, indent=1)
    row_item(32, "Resultados del Ejercicio",    0, indent=1)
    row_item(33, "TOTAL PATRIMONIO",            "=SUM(B31:B32)", bold=True, bg=GREEN)

    row_item(35, "TOTAL PASIVOS + PATRIMONIO",  "=B28+B33", bold=True, bg=LIGHT_BLUE)

    # Balance check
    check_row = 37
    ws.cell(row=check_row, column=1,
            value="✔ Balance (debe ser 0):").font = normal_font(bold=True)
    c = ws.cell(row=check_row, column=2, value="=B17-B35")
    c.number_format = '#,##0'
    c.font = normal_font(bold=True)


# ── Sheet 7 – IVA ───────────────────────────────────────────────────────────

def build_iva(wb):
    ws = wb.create_sheet("Registro IVA")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "Fecha", "Tipo\n(Venta/Compra)", "RUT Contraparte",
        "Razón Social", "N° Documento", "Tipo Doc.",
        "Monto Neto (CLP)", "IVA 19% (CLP)", "Total (CLP)"
    ]
    apply_header_row(ws, 1, headers)
    set_col_widths(ws, [14, 16, 16, 30, 16, 14, 18, 16, 16])

    # Example rows
    examples = [
        ["2024-01-05", "Compra",  "76.XXX.XXX-X", "Proveedor S.A.",  "0001234", "Factura", 500000, 95000,  595000],
        ["2024-01-10", "Venta",   "12.XXX.XXX-X", "Cliente Ltda.",   "0000001", "Factura", 300000, 57000,  357000],
    ]
    for i, row_data in enumerate(examples, start=2):
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        apply_data_row(ws, i, row_data, bg=bg)
        for col in (7, 8, 9):
            ws.cell(row=i, column=col).number_format = '#,##0'

    # Totals
    t = len(examples) + 2
    ws.cell(row=t, column=6, value="TOTALES").font = normal_font(bold=True)
    ws.cell(row=t, column=6).fill = fill(ORANGE)
    ws.cell(row=t, column=6).border = thin_border()
    for col in (7, 8, 9):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=t, column=col,
                       value=f"=SUM({col_letter}2:{col_letter}{t-1})")
        cell.number_format = '#,##0'
        cell.font = normal_font(bold=True)
        cell.fill = fill(ORANGE)
        cell.border = thin_border()

    # IVA summary box
    ws.cell(row=t + 2, column=1, value="RESUMEN IVA DEL PERÍODO").font = header_font(color=DARK_BLUE)
    ws.cell(row=t + 3, column=1, value="Débito Fiscal (IVA Ventas)").font = normal_font()
    ws.cell(row=t + 3, column=2, value=0).number_format = '#,##0'
    ws.cell(row=t + 4, column=1, value="Crédito Fiscal (IVA Compras)").font = normal_font()
    ws.cell(row=t + 4, column=2, value=0).number_format = '#,##0'
    ws.cell(row=t + 5, column=1, value="IVA A PAGAR / RECUPERAR").font = normal_font(bold=True)
    cell = ws.cell(row=t + 5, column=2,
                   value=f"=B{t+3}-B{t+4}")
    cell.number_format = '#,##0'
    cell.font = normal_font(bold=True)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    build_portada(wb)
    build_plan_cuentas(wb)
    build_libro_diario(wb)
    build_ingresos_egresos(wb)
    build_estado_resultados(wb)
    build_balance(wb)
    build_iva(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Template saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
